#CÂU1

#File XuLyFile.py
def LuuFile(path, data):
    file = open(path, 'a', encoding='utf-8')
    file.writelines(data)
    file.writelines("\n")
    file.close()

def DocFile(path):
    arrProduct = []
    file = open(path, 'r', encoding='utf-8')
    for line in file:
        data = line.strip()
        arr = data.split(";")
        arrProduct.append(arr)
    file.close()
    return arrProduct

#File TestLuuFile.py
masp = input("Nhập mã sản phẩm: ")
tensp = input("Nhập tên sản phẩm: ")
dongia = int(input("Nhập đơn giá: "))
line = masp + ";" + tensp + ";" + str(dongia)

LuuFile("database.txt", line)

#File TestDocFile.py
dssp = DocFile("database.txt")
def XuatSanPham(dssp):
    for row in dssp:
        for element in row:
            print(element, end="\t  Untitled1:33 - Chuong7.py:33")
        print()
    print()
XuatSanPham(dssp)

def  SortSp(dssp):
    for i in range(len(dssp)):
        for j in range(len(dssp)):
            a = dssp[i]
            b = dssp[j]
            if (a[2]) > (b[2]):
                dssp[i] = b
                dssp[j] = a
SortSp(dssp)
print("Sản phẩm sau khi sắp xếp giá:  Untitled1:47 - Chuong7.py:47")
XuatSanPham(dssp)
#CÂU2

#Xử lí file
def LuuFile(path,data):
    file=open(path,'a',encoding='utf-8')
    file.writelines(data)
    file.writelines("\n")
    file.close()
def DocFile(path):
    arrSo=[]
    file=open(path,'r',encoding='utf-8')
    for line in file:
        data=line.strip()
        arr=data.split(',')
        arrSo.append(arr)
    file.close()    
    return arrSo
    
#Test lưu file
LuuFile("csdl_so.txt","-5,4,7,9,3,20")
LuuFile("csdl_so.txt","5,-4,37,-19,24,-21")
LuuFile("csdl_so.txt","15,9,0,-38,-3,15")
LuuFile("csdl_so.txt","5,-4,77,-9,3,-7")
LuuFile("csdl_so.txt","55,44,27")
LuuFile("csdl_so.txt","-50,26")
#Test đọc file
arrSo=DocFile("csdl_so.txt")
print(arrSo)
def XuatSoAmTrenMoiDong(arrSo):
    for row in arrSo:
        for element in row:
            number=int(element)
            if number<0:
                print(number,end='\t  Untitled1:82 - Chuong7.py:82')
        print()
print("Các số âm trên mỗi dòng:  Untitled1:84 - Chuong7.py:84")
XuatSoAmTrenMoiDong(arrSo)
#CÂU3
from xml.dom.minidom import parse
import xml.dom.minidom
# Mở file xml bằng minidom parser
DOMTree = xml.dom.minidom.parse("employees.xml")
collection = DOMTree.documentElement
# Lấy tất cả tag là employee
employees = collection.getElementsByTagName("employee")
# Duyệt vòng lặp để lấy toàn bộ dữ liệu ra
for employee in employees:
 tag_id = employee.getElementsByTagName('id')[0]
 id=tag_id.childNodes[0].data
 tag_name = employee.getElementsByTagName('name')[0]
 name=tag_name.childNodes[0].data
 print(id,'\t  Untitled1:100 - Chuong7.py:100',name)
#CÂU4
import json
jsonString = '{ "ma":"nv1", "age":50, "ten":"Trần Duy Thanh"}'
dataObject=json.loads(jsonString)
print(dataObject)
print("Mã=  Untitled1:106 - Chuong7.py:106",dataObject["ma"])
print("Tên=  Untitled1:107 - Chuong7.py:107",dataObject["age"])
print("Tuổi=  Untitled1:108 - Chuong7.py:108",dataObject["ten"])
#CÂU5
import json
pythonObject = {
    "Ten":"Tran Duy Thanh",
    "Tuoi":50,
    "Ma":"nv1"
}

jsonString = json.dumps(pythonObject)
print(jsonString)
#CÂU6
import csv
with open('datacsv.csv', newline="") as f:
    reader = csv.reader(f, delimiter=';', quoting=csv.QUOTE_NONE)
    for row in reader:
        print(row[0],"\t  Untitled1:124 - Chuong7.py:124",row[1])
#CÂU7
import xlsxwriter
# Tạo một file excel cùng 1 sheet
workbook = xlsxwriter.Workbook('demo.xlsx')
worksheet = workbook.add_worksheet()
# thiết lập các cột cho file
worksheet.set_column('A:A', 5)
worksheet.set_column('B:B', 15)
worksheet.set_column('C:C', 20)
worksheet.set_column('D:D', 15)
worksheet.set_column('E:E', 15)
# định dạng tiêu đề cột in đậm
bold = workbook.add_format({'bold': True})
# thêm dòng tiêu đề và định dạng in đậm

worksheet.write('A1', 'STT',bold)
worksheet.write('B1', 'MÃ SẢN PHẨM',bold)
worksheet.write('C1', 'TÊN SẢN PHẨM',bold)
worksheet.write('D1', 'SỐ LƯỢNG',bold)
worksheet.write('E1', 'ĐƠN GIÁ',bold)
#thêm một dòng dữ liệu
worksheet.write('A2',1)
worksheet.write('B2','SP1')
worksheet.write('C2', 'Coca')
worksheet.write('D2', '15')
worksheet.write('E2', '15000')
#thêm một dòng dữ liệu
worksheet.write('A3',2)
worksheet.write('B3','SP2')
worksheet.write('C3', 'Pepsi')
worksheet.write('D3', '20')
worksheet.write('E3', '18000')
#Chèn Logo vào
worksheet.insert_image('B5', 'logo_UEL.png')
workbook.close()
#CÂU8
from openpyxl import load_workbook
wb = load_workbook('demo.xlsx')
print (wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for row in ws.values:
    for value in row:
        print(value,"\t  Untitled1:167 - Chuong7.py:167",end='')
        print("")
#CÂU9

import os

class Category:
    def __init__(self, ma, ten):
        self.ma = ma
        self.ten = ten

class Product:
    def __init__(self, ma, ten, dongia, madm):
        self.ma = ma
        self.ten = ten
        self.dongia = float(dongia)
        self.madm = madm

class ProductManager:
    def __init__(self, file_path="sanpham.txt"):
        self.file_path = file_path
        self.products = []
        self.load_data()

    def load_data(self):
        if not os.path.exists(self.file_path):
            return
        with open(self.file_path, "r", encoding="utf-8") as f:
            for line in f:
                ma, ten, dongia, madm = line.strip().split("|")
                self.products.append(Product(ma, ten, dongia, madm))

    def save_data(self):
        with open(self.file_path, "w", encoding="utf-8") as f:
            for p in self.products:
                f.write(f"{p.ma}|{p.ten}|{p.dongia}|{p.madm}\n")

    def add_product(self, p):
        self.products.append(p)
        self.save_data()

    def update_product(self, ma, ten=None, dongia=None, madm=None):
        for p in self.products:
            if p.ma == ma:
                if ten: p.ten = ten
                if dongia: p.dongia = float(dongia)
                if madm: p.madm = madm
                self.save_data()
                return True
        return False

    def delete_product(self, ma):
        self.products = [p for p in self.products if p.ma != ma]
        self.save_data()

    def search(self, keyword):
        return [p for p in self.products if keyword.lower() in p.ten.lower()]

    def sort_by_price(self):
        self.products.sort(key=lambda x: x.dongia)
        self.save_data()

# --- Demo ---
if __name__ == "__main__":
    manager = ProductManager()

    # Thêm dữ liệu mẫu
    manager.add_product(Product("SP01", "Bút bi", 5000, "DM01"))
    manager.add_product(Product("SP02", "Vở kẻ ngang", 12000, "DM01"))

    print("Danh sách sản phẩm:  Untitled1:237 - Chuong7.py:237")
    for p in manager.products:
        print(f"{p.ma}  {p.ten}  {p.dongia}  {p.madm}  Untitled1:239 - Chuong7.py:239")

    print("\nTìm 'Bút':  Untitled1:241 - Chuong7.py:241", [p.ten for p in manager.search("Bút")])
#CÂU10
import json

# ================== TẠO DỮ LIỆU ==================
def tao_lop(ma_lop, ten_lop):
    return {"ma_lop": ma_lop, "ten_lop": ten_lop}

def tao_sinhvien(ma_sv, ten_sv, nam_sinh, ma_lop):
    return {"ma_sv": ma_sv, "ten_sv": ten_sv, "nam_sinh": nam_sinh, "ma_lop": ma_lop}

# ================== QUẢN LÝ ==================
lops = []
sinhviens = []

# ----- LỚP -----
def them_lop(ma, ten):
    lops.append(tao_lop(ma, ten))

def tim_lop(ma):
    for lop in lops:
        if lop["ma_lop"] == ma:
            return lop
    return None

# ----- SINH VIÊN -----
def them_sinhvien(ma_sv, ten_sv, nam_sinh, ma_lop):
    if not tim_lop(ma_lop):
        print("Lớp không tồn tại!  Untitled1:269 - Chuong7.py:269")
        return
    sinhviens.append(tao_sinhvien(ma_sv, ten_sv, nam_sinh, ma_lop))

def sua_sinhvien(ma_sv, ten=None, nam_sinh=None):
    for sv in sinhviens:
        if sv["ma_sv"] == ma_sv:
            if ten: sv["ten_sv"] = ten
            if nam_sinh: sv["nam_sinh"] = nam_sinh
            print("Đã sửa sinh viên  Untitled1:278 - Chuong7.py:278")
            return
    print("Không tìm thấy sinh viên  Untitled1:280 - Chuong7.py:280")

def xoa_sinhvien(ma_sv):
    global sinhviens
    truoc = len(sinhviens)
    sinhviens = [sv for sv in sinhviens if sv["ma_sv"] != ma_sv]
    if len(sinhviens) < truoc:
        print("Đã xóa sinh viên  Untitled1:287 - Chuong7.py:287")
    else:
        print("Không tìm thấy sinh viên  Untitled1:289 - Chuong7.py:289")

def tim_sinhvien(keyword):
    kq = []
    for sv in sinhviens:
        if keyword.lower() in sv["ten_sv"].lower():
            kq.append(sv)
    return kq

def sapxep_sinhvien(by="ten"):
    if by == "ten":
        sinhviens.sort(key=lambda sv: sv["ten_sv"])
    elif by == "namsinh":
        sinhviens.sort(key=lambda sv: sv["nam_sinh"])

# ================== FILE ==================
def luu_file(filename="sinhvien.json"):
    data = {"lops": lops, "sinhviens": sinhviens}
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("Đã lưu file  Untitled1:309 - Chuong7.py:309")

def doc_file(filename="sinhvien.json"):
    global lops, sinhviens
    try:
        with open(filename, "r", encoding="utf-8") as f:
            data = json.load(f)
            lops = data["lops"]
            sinhviens = data["sinhviens"]
        print("Đã đọc file  Untitled1:318 - Chuong7.py:318")
    except FileNotFoundError:
        print("Chưa có file dữ liệu  Untitled1:320 - Chuong7.py:320")

# ================== DEMO ==================
if __name__ == "__main__":
    # Thêm lớp
    them_lop("CNTT11", "Công nghệ thông tin 11")
    them_lop("CNTT25", "Công nghệ thông tin 25")

    # Thêm sinh viên
    them_sinhvien("SV01", "Nguyễn Văn Chí Hào", 2005, "CNTT11")
    them_sinhvien("SV02", "Trần Hiếu Hậu", 2004, "CNTT25")
    them_sinhvien("SV03", "Trương Quốc Khải", 2006, "CNTT11")

    # Tìm kiếm
    print("\n Tìm sinh viên 'Nguyễn':  Untitled1:334 - Chuong7.py:334")
    for sv in tim_sinhvien("Nguyễn"):
        print(sv)

    # Sắp xếp
    print("\nDanh sách sinh viên sắp xếp theo năm sinh:  Untitled1:339 - Chuong7.py:339")
    sapxep_sinhvien(by="namsinh")
    for sv in sinhviens:
        print(sv)

    # Lưu file
    luu_file()

    # Đọc lại file
    doc_file()
    print("\nDanh sách sinh viên sau khi đọc file:  Untitled1:349 - Chuong7.py:349")
    for sv in sinhviens:
        print(sv)
#CÂU11

# employee_manager_excel.py
from openpyxl import Workbook, load_workbook
import os

class NhanVien:
    def __init__(self, ma, ten, tuoi):
        self.ma = ma
        self.ten = ten
        self.tuoi = int(tuoi)

class QuanLyNhanVien:
    def __init__(self, file_path="nhanvien.xlsx"):
        self.file_path = file_path
        self.ds_nv = []

    # Ghi dữ liệu vào file Excel
    def luu_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVien"

        # Tiêu đề cột
        ws.append(["STT", "Mã", "Tên", "Tuổi"])

        # Ghi từng nhân viên
        for i, nv in enumerate(self.ds_nv, start=1):
            ws.append([i, nv.ma, nv.ten, nv.tuoi])

        wb.save(self.file_path)
        print(f"Đã lưu dữ liệu vào '{self.file_path}'  Untitled1:383 - Chuong7.py:383")

    # Đọc danh sách nhân viên từ file Excel
    def doc_file(self):
        if not os.path.exists(self.file_path):
            print("File chưa tồn tại!  Untitled1:388 - Chuong7.py:388")
            return

        wb = load_workbook(self.file_path)
        ws = wb.active

        self.ds_nv.clear()
        for row in ws.iter_rows(min_row=2, values_only=True):
            _, ma, ten, tuoi = row
            self.ds_nv.append(NhanVien(ma, ten, tuoi))
        print("Đã đọc dữ liệu từ file Excel!  Untitled1:398 - Chuong7.py:398")

    # Thêm nhân viên
    def them_nv(self, nv):
        self.ds_nv.append(nv)

    # Sắp xếp theo tuổi tăng dần
    def sap_xep_tuoi(self):
        self.ds_nv.sort(key=lambda nv: nv.tuoi)
        print("Đã sắp xếp nhân viên theo tuổi tăng dần!  Untitled1:407 - Chuong7.py:407")

    # Hiển thị danh sách nhân viên
    def hien_thi(self):
        print(f"\n{'Mã':<10}{'Tên':<15}{'Tuổi':<5}  Untitled1:411 - Chuong7.py:411")
        print("" * 30)
        for nv in self.ds_nv:
            print(f"{nv.ma:<10}{nv.ten:<15}{nv.tuoi:<5}  Untitled1:414 - Chuong7.py:414")

# --- Demo ---
if __name__ == "__main__":
    ql = QuanLyNhanVien()

    # Thêm dữ liệu mẫu
    ql.them_nv(NhanVien("NV1", "Khải", 21))
    ql.them_nv(NhanVien("NV2", "Huy", 21))
    ql.them_nv(NhanVien("NV3", "Hậu", 20))
    ql.them_nv(NhanVien("NV4", "Hào", 21))
    ql.them_nv(NhanVien("NV5", "Đào", 20))
    ql.them_nv(NhanVien("NV6", "Danh", 240))

    # Lưu file
    ql.luu_file()

    # Đọc lại file
    ql.doc_file()

    # Sắp xếp theo tuổi
    ql.sap_xep_tuoi()
    ql.hien_thi()
#CÂU12
import random
import csv


# Hàm lưu file CSV
def luu_csv(filename="data.csv"):
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        for _ in range(10):  # 10 dòng
            row = [random.randint(1, 100) for _ in range(10)]  # 10 số ngẫu nhiên
            writer.writerow(row)
    print(f"💾 Đã lưu file {filename}  Untitled1:449 - Chuong7.py:449")


# Hàm đọc file CSV và tính tổng từng dòng
def doc_csv(filename="data.csv"):
    with open(filename, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=";")
        for i, row in enumerate(reader, start=1):
            numbers = [int(x) for x in row if x.strip() != ""]
            tong = sum(numbers)
            print(f"Dòng {i}: {row} ➝ Tổng = {tong}  Untitled1:459 - Chuong7.py:459")


# ===================== DEMO =====================

if __name__ == "__main__":
    luu_csv("data.csv")
    doc_csv("data.csv")
#CÂU13

import xml.etree.ElementTree as ET
from collections import defaultdict

# ----------- Đọc file XML -----------
def doc_nhom(file_nhom):
    tree = ET.parse(file_nhom)
    root = tree.getroot()
    ds_nhom = []
    for nhom in root.findall("nhom"):
        ma = nhom.find("ma").text
        ten = nhom.find("ten").text
        ds_nhom.append({"ma": ma, "ten": ten})
    return ds_nhom

def doc_thietbi(file_tb):
    tree = ET.parse(file_tb)
    root = tree.getroot()
    ds_tb = []
    for tb in root.findall("thietbi"):
        manhom = tb.get("manhom")
        ma = tb.find("ma").text
        ten = tb.find("ten").text
        ds_tb.append({"manhom": manhom, "ma": ma, "ten": ten})
    return ds_tb

# ----------- Hiển thị danh sách -----------
def hien_thi_nhom(ds_nhom):
    print("\nDANH SÁCH NHÓM THIẾT BỊ:  Untitled1:496 - Chuong7.py:496")
    for n in ds_nhom:
        print(f"{n['ma']}: {n['ten']}  Untitled1:498 - Chuong7.py:498")

def hien_thi_thietbi(ds_tb):
    print("\nDANH SÁCH TOÀN BỘ THIẾT BỊ:  Untitled1:501 - Chuong7.py:501")
    for tb in ds_tb:
        print(f"{tb['ma']:<5} | {tb['ten']:<15} | Nhóm: {tb['manhom']}  Untitled1:503 - Chuong7.py:503")

# ----------- Lọc thiết bị theo nhóm -----------
def loc_theo_nhom(ds_tb, ma_nhom):
    kq = [tb for tb in ds_tb if tb["manhom"] == ma_nhom]
    return kq

# ----------- Nhóm có nhiều thiết bị nhất -----------
def nhom_nhieu_thietbi(ds_tb):
    dem = defaultdict(int)
    for tb in ds_tb:
        dem[tb["manhom"]] += 1
    nhom_max = max(dem, key=dem.get)
    return nhom_max, dem[nhom_max]

# ----------- Chương trình chính -----------
if __name__ == "__main__":
    file_nhom = "nhomthietbi.xml"
    file_tb = "thietbi.xml"

    ds_nhom = doc_nhom(file_nhom)
    ds_tb = doc_thietbi(file_tb)

    hien_thi_nhom(ds_nhom)
    hien_thi_thietbi(ds_tb)

    ma_nhom = input("\nNhập mã nhóm cần lọc (vd: n1): ").strip()
    kq = loc_theo_nhom(ds_tb, ma_nhom)
    print(f"\nThiết bị thuộc nhóm {ma_nhom}:  Untitled1:531 - Chuong7.py:531")
    for tb in kq:
        print(f"{tb['ma']}: {tb['ten']}  Untitled1:533 - Chuong7.py:533")

    manhom_max, so_luong = nhom_nhieu_thietbi(ds_tb)
    print(f"\nNhóm có nhiều thiết bị nhất: {manhom_max} ({so_luong} thiết bị)  Untitled1:536 - Chuong7.py:536")
